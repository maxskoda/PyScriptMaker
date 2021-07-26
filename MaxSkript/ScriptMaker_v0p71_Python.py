# -*- coding: utf-8 -*-
"""
Created on Thu May 12 09:09:50 2020

@author: Maximilian Skoda
"""

# Standard import
import importlib
# import qdarkstyle
import json
# import zmq
import logging
import os.path

from PyQt5 import QtGui, QtCore, QtWidgets
from PyQt5.QtCore import (QModelIndex, Qt)
from PyQt5.QtGui import QIntValidator
from PyQt5.QtWidgets import (QLabel, QMessageBox, QApplication, qApp, QFileDialog, QSizePolicy)
from openpyxl import load_workbook

from MaxSkript.sampletable import SpinBoxDelegate
from MaxSkript.treeview import Tree

# import custom 'actions' file
# import MuonActions #NRActions

# define name string for dynamic import of action classes:

myActions = "Actions.NRActions_Python"

NRActions = __import__('Actions.NRActions_Python', globals(), locals(), ['NRActions_Python'], 0)

# HORIZONTAL_HEADERS = ("Action", "Parameters", "Ok", "Row", "Action duration / min")

logging.basicConfig(level=logging.DEBUG)
countRate = 40


class App(QtWidgets.QWidget):
    def __init__(self, parent=None):
        super(App, self).__init__(parent)

        global NRActions
        global myActions
        self.view = Tree(self)

        rows = self.view.model.rowCount()
        for i in range(rows):
            self.view.show_summary(self.view.model.index(i, 0))

        # set the font
        myFont = "Verdana"  # Consolas ok
        font = QtGui.QFont(myFont, 10)
        self.view.setFont(font)

        self.dataGroupBox = QtWidgets.QGroupBox("NR Script")
        self.dataGroupBox.setFont(font)
        dataLayout = QtWidgets.QHBoxLayout()

        # dataLayout.addWidget(actionList)
        dataLayout.addWidget(self.view)

        self.dataGroupBox.setLayout(dataLayout)

        self.mainLayout = QtWidgets.QVBoxLayout()

        self.instrumentSelector = QtWidgets.QComboBox()
        self.instrumentSelector.addItems(NRActions.instruments)
        self.instrumentSelector.setFont(QtGui.QFont(myFont, 12, QtGui.QFont.Black))

        buttonLayout = QtWidgets.QHBoxLayout()
        buttonLayout.addWidget(self.instrumentSelector)
        self.printTreeButton = QtWidgets.QPushButton("Output script (Ctrl+p)", self)
        self.printTreeButton.setFont(QtGui.QFont(myFont, 12, QtGui.QFont.Black))
        self.printTreeButton.setShortcut('Ctrl+p')
        self.printTreeButton.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self.printTreeButton.clicked.connect(self.on_print_tree)
        buttonLayout.addWidget(self.printTreeButton)

        self.playButton = QtWidgets.QPushButton("", self)
        self.playButton.setFont(QtGui.QFont(myFont, 12, QtGui.QFont.Black))
        self.playButton.setShortcut('Ctrl+r')
        self.playButton.setIcon(QtGui.QIcon('MaxSkript/Icons/play_icon.png'))
        self.playButton.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self.playButton.clicked.connect(self.play_script)
        buttonLayout.addWidget(self.playButton)

        self.pauseButton = QtWidgets.QPushButton("", self)
        self.pauseButton.setFont(QtGui.QFont(myFont, 12, QtGui.QFont.Black))
        self.pauseButton.setShortcut('Ctrl+/')
        self.pauseButton.setIcon(QtGui.QIcon('MaxSkript/Icons/pause_icon.png'))
        self.pauseButton.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self.pauseButton.clicked.connect(self.pause_script)
        buttonLayout.addWidget(self.pauseButton)

        self.stopButton = QtWidgets.QPushButton("", self)
        self.stopButton.setFont(QtGui.QFont(myFont, 12, QtGui.QFont.Black))
        self.stopButton.setShortcut('Ctrl+.')
        self.stopButton.setIcon(QtGui.QIcon('MaxSkript/Icons/stop_icon.png'))
        self.stopButton.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self.stopButton.clicked.connect(self.stop_script)
        buttonLayout.addWidget(self.stopButton)

        self.printSamplesButton = QtWidgets.QPushButton("Print samples")
        self.printSamplesButton.clicked.connect(self.on_print_samples)
        # buttonLayout.addWidget(self.printSamplesButton)
        self.timeLabel = QLabel("Duration:~ ")
        self.timeLabel.setAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)
        self.timeLabel.setFont(QtGui.QFont(myFont, 12, QtGui.QFont.Black))
        self.runTime = QLabel("00h 00min")
        self.runTime.setFont(QtGui.QFont(myFont, 12, QtGui.QFont.Black))
        buttonLayout.addWidget(self.timeLabel)
        buttonLayout.addWidget(self.runTime)

        self.actionsLabel = QtWidgets.QLabel("Actions file: ")
        self.actionsEdit = QtWidgets.QLineEdit()
        self.actionsOpenButton = QtWidgets.QPushButton("...")
        self.actionsOpenButton.clicked.connect(self.on_refresh_actions)

        self.fileLabel = QtWidgets.QLabel("Script file: ")
        self.fileEdit = QtWidgets.QLineEdit()
        self.fileOpenButton = QtWidgets.QPushButton("...")
        self.fileOpenButton.clicked.connect(self.on_open_file)

        fileLayout = QtWidgets.QHBoxLayout()
        fileLayout.addWidget(self.actionsLabel)
        fileLayout.addWidget(self.actionsEdit)
        fileLayout.addWidget(self.actionsOpenButton)
        fileLayout.addWidget(self.fileLabel)
        fileLayout.addWidget(self.fileEdit)
        fileLayout.addWidget(self.fileOpenButton)

        # button to collapse all tree items
        self.collapseButton = QtWidgets.QPushButton("", self)
        self.collapseButton.setShortcut('Ctrl+-')
        self.collapseButton.setToolTip("Ctrl+-")
        self.collapseButton.setIcon(QtGui.QIcon('MaxSkript/Icons/collapse_all_icon.png'))
        self.collapseButton.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self.collapseButton.clicked.connect(self.collapse_all)

        # button to expand all tree items
        self.expandButton = QtWidgets.QPushButton("", self)
        self.expandButton.setShortcut('Ctrl++')
        self.expandButton.setToolTip("Ctrl++")
        self.expandButton.setIcon(QtGui.QIcon('MaxSkript/Icons/expand_all_icon.png'))
        self.expandButton.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self.expandButton.clicked.connect(self.expand_all)

        # Beam current override
        self.beam_current_label = QtWidgets.QLabel("Beam current:", self)
        self.beam_current_label.setFont(QtGui.QFont(myFont, 12, QtGui.QFont.Black))
        self.beam_current_edit = QtWidgets.QLineEdit("", self)
        self.onlyInt = QIntValidator()
        self.beam_current_edit.setValidator(self.onlyInt)
        self.beam_current_edit.setFixedWidth(120)
        self.beam_current_edit.setFont(QtGui.QFont(myFont, 12, QtGui.QFont.Black))
        self.beam_current_edit.editingFinished.connect(self.set_uamps)
        self.uamp_label = QtWidgets.QLabel("\u03BCA", self)
        self.uamp_label.setFont(QtGui.QFont(myFont, 12, QtGui.QFont.Black))

        collapseLayout = QtWidgets.QHBoxLayout()
        collapseLayout.addWidget(self.collapseButton)
        collapseLayout.addWidget(self.expandButton)
        collapseLayout.addWidget(self.beam_current_label)
        collapseLayout.addWidget(self.beam_current_edit)
        collapseLayout.addWidget(self.uamp_label)

        self.mainLayout.addLayout(fileLayout)
        self.mainLayout.addLayout(buttonLayout)
        self.mainLayout.addLayout(collapseLayout)

        self.view.resizeColumnToContents(0)
        self.view.resizeColumnToContents(2)
        self.view.header().resizeSection(2, 5)

        self.show()

    ###########
    def play_script(self):
        self.view.model.getRootItem(0, 0).setBackground(QtGui.QBrush(Qt.BDiagPattern))
        self.view.model.getRootItem(0, 1).setBackground(QtGui.QBrush(Qt.BDiagPattern))
        self.playButton.setDisabled(1)
        self.pauseButton.setEnabled(1)

    def pause_script(self):
        self.view.model.getRootItem(0, 0).setBackground(QtGui.QBrush(Qt.BDiagPattern))
        self.view.model.getRootItem(0, 1).setBackground(QtGui.QBrush(Qt.BDiagPattern))
        if not self.playButton.isEnabled() and self.pauseButton.isEnabled():
            self.pauseButton.setDisabled(1)
            self.playButton.setEnabled(1)

    def stop_script(self):
        self.view.model.getRootItem(0, 0).setBackground(QtGui.QBrush(Qt.NoBrush))
        self.view.model.getRootItem(0, 1).setBackground(QtGui.QBrush(Qt.NoBrush))
        self.playButton.setDisabled(0)
        self.pauseButton.setDisabled(0)

    def collapse_all(self):
        self.view.collapseAll()

    def expand_all(self):
        self.view.expandAll()

    def set_uamps(self):
        global countRate
        if self.beam_current_edit:
            countRate = self.beam_current_edit.text()
            # print(countRate)

    def on_refresh_actions(self):
        # print(self.actionsEdit.text())
        global myActions
        global NRActions
        # del myActions
        NRActions = importlib.import_module(self.actionsEdit.text())
        importlib.reload(NRActions)
        myActions = self.actionsEdit.text()

    def on_open_file(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        fileName, _ = QFileDialog.getSaveFileName(self, "Open Script File...", "", \
                                                  "Python Files (*.py);;All Files (*);;OpenGenie Files (*.gcl)",
                                                  options=options)

        try:
            with open(fileName) as f:
                self.fileEdit.setText(fileName)
                return fileName
        except IOError:
            if fileName:
                print("Created new file " + fileName)
                self.fileEdit.setText(fileName)
                f = open(fileName, "w")
                f.close()

        # if fileName:
        #     self.fileEdit.setText(fileName)
        #     return fileName

    def on_print_tree(self):
        self.view.update_summary()
        if os.path.isfile(self.fileEdit.text()):
            f = open(self.fileEdit.text(), "w+")
        else:
            f = self.on_open_file()
            if os.path.isfile(self.fileEdit.text()):
                f = open(self.fileEdit.text(), "w+")
            else:
                return

        # NEED TO REVISIT THIS
        args = []
        for col in range(self.view.tableModel.columnCount(QModelIndex)):
            args.append(0)
            # if self.sampleTable.item(0,col).checkState()  == Qt.Checked:
            #     args.append(1)
            # else:
            #     args.append(0)
        f.write(NRActions.writeHeader(self.on_print_samples(), args))

        for row in range(self.view.model.rowCount()):
            it = self.view.model.item(row)
            MyClass = getattr(importlib.import_module(myActions), it.text())
            # Instantiate the class (pass arguments to the constructor, if needed)
            tempAction = MyClass()
            tempAction.makeAction(it)
            # Check if sample entry exists for Action:
            if it.child(0, 0).text() == "Sample":
                sampleNumber = self.view.delegate.items.index(it.child(0, 1).text()) + 1
            else:
                sampleNumber = -1
            f.write(tempAction.stringLine(sampleNumber) + "\n")

        f.write(NRActions.writeFooter(self.on_print_samples()))
        f.close()

    def on_print_samples(self):
        samples = []
        # print(self.view.tableModel.getData())
        for row in range(self.view.tableModel.rowCount(QModelIndex)):
            # print(self.view.tableModel.getData()[row][0])
            # print(self.view.tableModel.headerData(0, Qt.Horizontal, Qt.DisplayRole))

            if self.view.tableModel.getData()[row][0] != "":
                sampleDict = {}
                for col in range(self.view.tableModel.columnCount(QModelIndex)):
                    sampleDict[self.view.tableModel.headerData(col, Qt.Horizontal, Qt.DisplayRole)] = \
                    self.view.tableModel.getData()[row][col]
                    # print(sampleDict)
                samples.append(sampleDict)
        return samples



class MyMainWindow(QtWidgets.QMainWindow):

    def __init__(self, parent=None):

        super(MyMainWindow, self).__init__(parent)
        self.form_widget = App(self)
        self.setCentralWidget(self.form_widget)

        openAct = QtWidgets.QAction(QtGui.QIcon('exit.png'), '&Open...', self)
        openAct.setShortcut('Ctrl+O')
        openAct.setStatusTip('Open script')
        openAct.triggered.connect(self.openScript)

        saveAct = QtWidgets.QAction(QtGui.QIcon('exit.png'), '&Save as...', self)
        saveAct.setShortcut('Ctrl+S')
        saveAct.setStatusTip('Save script')
        saveAct.triggered.connect(self.saveScript)

        exitAct = QtWidgets.QAction(QtGui.QIcon('exit.png'), '&Exit', self)
        exitAct.setShortcut('Ctrl+Q')
        exitAct.setStatusTip('Exit application')
        exitAct.triggered.connect(qApp.quit)

        change_subtitle_Act = QtWidgets.QAction(QtGui.QIcon('exit.png'), 'E&dit subtitles', self)
        change_subtitle_Act.setShortcut('Ctrl+E')
        change_subtitle_Act.setStatusTip('Edit subtitles')
        change_subtitle_Act.triggered.connect(self.edit_subtitles)

        self.statusBar()

        menubar = self.menuBar()
        fileMenu = menubar.addMenu('&File')
        fileMenu.addAction(openAct)
        fileMenu.addAction(saveAct)
        fileMenu.addAction(exitAct)

        editMenu = menubar.addMenu('&Edit')
        editMenu.addAction(change_subtitle_Act)

        QtWidgets.qApp.installEventFilter(self)
        ################################
        # Sample Table - this is NR specific and should be moved to NRActions.py!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
        ################################
        self.NRSamples = QtWidgets.QDockWidget('Sample table', self)
        myFont = "Verdana"
        self.NRSamples.setFont(QtGui.QFont(myFont, 12, QtGui.QFont.Black))
        self.sampleTableGroupBox = QtWidgets.QGroupBox("Sample table")
        # self.sampleTableGroupBox.setFont(font)
        sampleLayout = QtWidgets.QHBoxLayout()

        self.sampleTable = QtWidgets.QTableView()

        # Create buttons:
        self.addLayerButton = QtWidgets.QPushButton("Add sample")
        self.addLayerButton.clicked.connect(self.on_addLayer)
        self.removeLayerButton = QtWidgets.QPushButton("Remove Sample(s)")
        self.removeLayerButton.clicked.connect(self.on_removeLayer)

        self.hbox_layers = QtWidgets.QHBoxLayout()

        self.hbox_layers.addWidget(self.addLayerButton)
        self.hbox_layers.addWidget(self.removeLayerButton)
        self.hbox_layers.addStretch(1)
        self.vbox_samples = QtWidgets.QVBoxLayout()

        # checkbox to hide table
        self.b = QtWidgets.QCheckBox("Hide/Show Sample Table", self)
        self.b.stateChanged.connect(self.click_box)

        self.vbox_samples.addWidget(self.b)
        self.vbox_samples.addLayout(self.hbox_layers)
        self.vbox_samples.addWidget(self.sampleTable)
        sampleLayout.addLayout(self.vbox_samples)
        self.sampleTableGroupBox.setLayout(sampleLayout)
        self.setLayout(self.form_widget.mainLayout)

        self.sampleTable.setAlternatingRowColors(True)
        self.sampleTable.setModel(self.form_widget.view.tableModel)
        self.sampleTable.resizeColumnsToContents()

        self.form_widget.mainLayout.addWidget(self.form_widget.dataGroupBox)
        self.form_widget.mainLayout.addWidget(self.b)
        # self.form_widget.mainLayout.addWidget(self.sampleTableGroupBox)
        self.NRSamples.setWidget(self.sampleTable)#sampleTableGroupBox)
        self.NRSamples.setFeatures(QtWidgets.QDockWidget.DockWidgetFloatable |
                 QtWidgets.QDockWidget.DockWidgetMovable)
        self.addDockWidget(Qt.BottomDockWidgetArea, self.NRSamples)

        self.NRSamples.setContentsMargins(20,20,20,20)
        self.form_widget.setLayout(self.form_widget.mainLayout)

        # setting style sheet to the NRSample dock widget
        # self.NRSamples.setStyleSheet(
        #                    "QDockWidget::title"
        #                    "{"
        #                    "background : darkgrey;"
        #                    "}"
        #                    "QDockWidget QWidget"
        #                    "{"
        #                    "border : 3px solid black;"
        #                    "}"
        #                    )
        style ="QDockWidget::Widget{padding - left: 30px;}"
        style += "QProgressBar {border: 2px solid grey; border-radius: 5px; text-align: center;}"
        style += "QProgressBar::chunk {background-color: #CD96CD; width: 10px; margin: 0.5px;}"

        self.setStyleSheet(style)
        ##################################


        # Open default file
        fileName = "MaxSkript/INTER_4Samples_2Contrasts_inject.json"  # "Muon_test.json"
        self.form_widget.view.model.populate(fileName)
        for col in range(self.form_widget.view.model.columnCount()):
            self.form_widget.view.resizeColumnToContents(col)

        with open(fileName) as json_file:
            data = json.load(json_file)
            """
            for row, samp in enumerate(data['Samples'], start=1):
                if len(samp):
                    for col in range(self.form_widget.sampleTable.columnCount()):
                        self.form_widget.sampleTable.item(row, col).setText(samp[col])
            """
        rows = self.form_widget.view.model.rowCount()
        for i in range(rows):
            if self.form_widget.view.model.item(i, 0).child(0, 0).text() == "Sample":
                self.form_widget.view.openPersistentEditor(self.form_widget.view.model.item(i, 0).child(0, 1).index())
            self.form_widget.view.show_summary(self.form_widget.view.model.index(i, 0))
        self.form_widget.view.resizeColumnToContents(0)
        self.form_widget.fileEdit.setText('MaxSkript/PyScript_test1.py')
        ############################
        self.setStyleSheet("""

            QTreeView {
            alternate-background-color: #e8f4fc;
            background: #f6fafb;
            }

        """)

    def click_box(self, state):
        if state == QtCore.Qt.Checked:
            self.NRSamples.hide()
            # self.addLayerButton.setHidden(True)
            # self.removeLayerButton.setHidden(True)
            # self.sampleTable.setHidden(True)
        else:
            self.NRSamples.show()
            # self.addLayerButton.setHidden(False)
            # self.removeLayerButton.setHidden(False)
            # self.sampleTable.setHidden(False)

    def on_addLayer(self):
        pass
        # data = self.view.tableModel.getData()
        # rowNo = len(data)
        # data.append(["S"+str(rowNo+1), "", "", "", "", "", "", "", ""])
        # self.view.tableModel = TableModel(data, self.view.headerLabels_sampTable)
        # self.sampleTable.setModel(self.view.tableModel)
        #
        # self.view.tableModel.rowsInserted.connect(self.view.delegate.updateCombo)

    def on_removeLayer(self):
        pass


    def openScript(self):
        print("Opening...")
        buttonReply = QMessageBox.question(self, 'MaxSkript', "All changes will be lost. Do you want to save?", \
                                           QMessageBox.Yes | QMessageBox.No | QMessageBox.Cancel, QMessageBox.Cancel)
        if buttonReply == QMessageBox.Yes:
            print('Yes clicked.')
            self.saveScript()
        if buttonReply == QMessageBox.No:
            print('No clicked.')
        if buttonReply == QMessageBox.Cancel:
            return

        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        fileName, _ = QFileDialog.getOpenFileName(self, "Open saved ScriptMaker state", "", \
                                                  "MaxScript Files (*.json);;"
                                                  "SANS Excel table (*.xlsx);;"
                                                  "All Files (*)", options=options)
        if fileName:
            if fileName.split(".")[1].upper() == "JSON":
                # delete all rows
                self.form_widget.view.model.removeRows(0, self.form_widget.view.model.rowCount())

                self.form_widget.view.model.populate(fileName)
                with open(fileName) as json_file:
                    data = json.load(json_file)
                    # print(len(data['Samples']))
                    # tableModel = TableModel(self.form_widget.view, data['Samples'], self.form_widget.view.headerLabels_sampTable)
                    self.form_widget.view.tableModel.layoutAboutToBeChanged.emit()
                    self.form_widget.view.sampleTable = data['Samples']
                    self.form_widget.view.update_sample_table()
                    self.form_widget.view.tableModel.layoutChanged.emit()

                rows = self.form_widget.view.model.rowCount()
                for i in range(rows):
                    self.form_widget.view.show_summary(self.form_widget.view.model.index(i, 0))
                self.form_widget.parent().setWindowTitle("Ma_xSkript - " + fileName + "[*]")

            else:  # SANS Excel table file
                # print(fileName.split(".")[1].upper())
                json_file_name = fileName.split(".")[0] + ".json"
                print(json_file_name)
                wb = load_workbook(filename=fileName)
                ws = wb.active
                wb_row = 2
                with open(json_file_name, "w+") as f:
                    outString = "{\"Samples\": ["
                    ## get defined samples from table:
                    data = self.form_widget.sampleTable.model().getData()

                    for row in data:
                        r = str([''.join(x) for x in row])
                        r = str(r).replace("\'", "\"")
                        outString += r + ",\n"
                    outString = outString[:-2] + "],\n"

                    outString += "\n\n\"Action\":[\n"
                    while ws.cell(row=wb_row, column=1).value:
                        do_trans = getattr(importlib.import_module(myActions), 'do_trans')
                        # Instantiate the class (pass arguments to the constructor, if needed)
                        tempAction = do_trans(title=str(ws.cell(row=wb_row, column=7).value),
                                              pos=str(ws.cell(row=wb_row, column=1).value),
                                              uamps=ws.cell(row=wb_row, column=2).value,
                                              thickness=ws.cell(row=wb_row, column=9).value)
                        outString += tempAction.makeJSON() + ","
                        del tempAction

                        do_sans = getattr(importlib.import_module(myActions), 'do_sans')
                        # Instantiate the class (pass arguments to the constructor, if needed)
                        tempAction = do_sans(title=str(ws.cell(row=wb_row, column=7).value) + '_' + \
                                                   str(ws.cell(row=wb_row, column=8).value),
                                             pos=str(ws.cell(row=wb_row, column=1).value),
                                             uamps=ws.cell(row=wb_row, column=4).value,
                                             thickness=ws.cell(row=wb_row, column=9).value)
                        outString += tempAction.makeJSON() + ","
                        del tempAction
                        wb_row += 1
                    outString = outString[:-1] + "]\n}"
                    f.write(outString)

                # delete all rows
                self.form_widget.view.model.removeRows(0, self.form_widget.view.model.rowCount())

                self.form_widget.view.model.populate(json_file_name)
                with open(json_file_name) as json_file:
                    data = json.load(json_file)
                    self.form_widget.view.tableModel.layoutAboutToBeChanged.emit()
                    self.form_widget.view.sampleTable = data['Samples']
                    self.form_widget.view.update_sample_table()
                    self.form_widget.view.tableModel.layoutChanged.emit()

    def saveScript(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        fileName, _ = QFileDialog.getSaveFileName(self, "Save ScriptMaker state...", "", \
                                                  "Ma_xScript Files (*.json);;All Files (*)", options=options)
        if fileName:
            f = open(fileName, "w+")
        else:
            return

        outString = "{\"Samples\": ["
        ## get defined samples from table:
        data = self.form_widget.sampleTable.model().getData()

        for row in data:
            r = str([''.join(x) for x in row])
            r = str(r).replace("\'", "\"")
            outString += r + ",\n"
        outString = outString[:-2] + "],\n"

        outString += "\n\n\"Action\":[\n"
        for row in range(self.form_widget.view.model.invisibleRootItem().rowCount()):
            it = self.form_widget.view.model.item(row)
            MyClass = getattr(importlib.import_module(myActions), it.text())
            # Instantiate the class (pass arguments to the constructor, if needed)
            tempAction = MyClass()
            tempAction.makeAction(it)
            outString += tempAction.makeJSON()
            if row < self.form_widget.view.model.rowCount() - 1:
                outString += ","
            del tempAction

        outString += "]\n}"
        f.write(outString)
        f.close()
        print("Saved.")

    def edit_subtitles(self):
        new_subtitle = self.getText()
        for model_index in self.form_widget.view.selectionModel().selectedRows():
            row = model_index.row()
            it = self.form_widget.view.model.item(row)
            if it.child(1, 0).text() == 'Subtitle':
                it.child(1, 1).setText(new_subtitle)
                self.form_widget.view.show_summary(model_index)

    def getText(self):
        text, okPressed = QtWidgets.QInputDialog.getText(self, "Change subtitles", "New subtitle:",
                                                         QtWidgets.QLineEdit.Normal, "")
        if okPressed and text != '':
            return text


def main():
    app = QApplication([])
    app.aboutToQuit.connect(app.deleteLater)
    # app.setStyle('Plastique')
    # app.setStyleSheet(qdarkstyle.load_stylesheet(qt_api='pyqt5'))
    foo = MyMainWindow()
    foo.resize(700, 800)
    foo.setWindowTitle("Ma_xSkript [*]")
    # foo.setWindowFilePath("fiel C:/Users/ktd43279/Documents/GitHub/PyScriptMaker/ScriptMaker_v0p71.py")
    foo.setWindowIcon(QtGui.QIcon('MaxSkript/Icons/squareCogs.gif'))

    # Force the style to be the same on all OSs:
    app.setStyle("Fusion")

    # # Now use a palette to switch to dark colors:
    # palette = QtGui.QPalette()
    # palette.setColor(QtGui.QPalette.Window, QtGui.QColor(53, 53, 53))
    # palette.setColor(QtGui.QPalette.WindowText, Qt.white)
    # app.setPalette(palette)

    foo.show()
    app.exec_()


if __name__ == '__main__':
    main()
