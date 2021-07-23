# -*- coding: utf-8 -*-
"""
Created on Thu May 12 09:09:50 2020

@author: Maximilian Skoda
"""

import qdarkstyle
import json
import os.path
import zmq
import logging

from PyQt5.QtCore import (QAbstractItemModel, QFile, QIODevice, pyqtSlot, QSortFilterProxyModel,
                          QModelIndex, Qt, QDataStream, QVariant)
from PyQt5 import QtGui, QtCore, QtWidgets
from PyQt5.QtGui import QKeySequence
from PyQt5.QtWidgets import (QWidget, QLabel, QMessageBox, QShortcut, QListWidgetItem, QStyleOptionProgressBar,
    QComboBox, QApplication, QTreeWidget, qApp, QFileDialog, QAbstractItemView, QStyledItemDelegate, QDoubleSpinBox,
                             QSpinBox, QLCDNumber, QSizePolicy)

from functools import partial
from openpyxl import load_workbook
# import numpy as np

# Standard import
import importlib

# import custom 'actions' file
# import MuonActions #SANSActions

# define name string for dynamic import of action classes:
myActions = "MARIActions"#"SANSActions_Python" #"MuonActions"

SANSActions = __import__('MARIActions') #__import__('MuonActions')

HORIZONTAL_HEADERS = ("Action", "Parameters", "Ok", "Row", "Action duration / min")

logging.basicConfig(level=logging.DEBUG)

class myStandardItemModel(QtGui.QStandardItemModel):
    def __init__(self, parent = None):
        super(myStandardItemModel, self).__init__(parent)

    def canDropMimeData(self, data, action, row, column, parent):
        print('can drop called on')
        print(parent.data())
        print("PARENT: ", parent.parent())
        return True

    def dropEvent(self, e):
        self.setText("Hallo")
    # def dropMimeData(self, data, action, row, column, parent):
    #     parent_name = parent.data()
    #     node_name = data.text()
    #     print("Dropped {} onto {}".format(node_name, parent_name))
    #     return True

    def flags(self, index):
        defaultFlags =  QtCore.Qt.ItemIsEnabled | super(myStandardItemModel, self).flags(index)
        
        if index.column() == 1:
            return defaultFlags | QtCore.Qt.ItemIsEditable
        
        if index.isValid():
            return defaultFlags | Qt.ItemIsDragEnabled# | Qt.ItemIsDropEnabled
        else:
            return defaultFlags | Qt.ItemIsDropEnabled

    def itemList(self, parent = QtCore.QModelIndex()):
        items = []
        print('itemlist')
        for row in range(self.rowCount(parent)):
            idx = self.index(row, 0, parent)
            items.append(self.data(idx))
            if self.hasChildren(idx):
                items.append(self.itemList(idx))
        return items

    def headerData(self, column, orientation, role):
        if orientation == QtCore.Qt.Horizontal and role == QtCore.Qt.DisplayRole:
            try:
                return QtCore.QVariant(HORIZONTAL_HEADERS[column])
            except IndexError:
                pass
        return QtCore.QVariant()

    def getItemText(self, row, col):
        return self.item(row,col).text()
    
    def getRootItem(self, row, col):
        return self.item(row, col)
    
    def setMyText(self, index, text):
        self.item(index.row(), index.column()).setText(text)

    def populate(self, file):
        if os.path.isfile(file):
            with open(file) as json_file:
                data = json.load(json_file)
            for p in data['Action']:
                parentItem = self.invisibleRootItem()
                parentItem.setFlags(parentItem.flags() & ~QtCore.Qt.ItemIsEditable & ~QtCore.Qt.ItemIsDropEnabled)
                
                if isinstance(p, dict):
                    for key in p.keys():
                        # rowNoItem = QtGui.QStandardItem("")
                        item2 = QtGui.QStandardItem("")
                        item = QtGui.QStandardItem(key)
                        itemCheck = QtGui.QStandardItem("")
                        lastItem = QtGui.QStandardItem("")
                        durationItem = QtGui.QStandardItem("||||||||||||||||||||||||||||||||||||||||")
                        #durationItem.setData(45, Qt.UserRole)
                        item.setFlags(item.flags() & ~QtCore.Qt.ItemIsEditable
                                                   & ~QtCore.Qt.ItemIsDropEnabled\
                                                   & ~QtCore.Qt.ItemIsDragEnabled)
                        itemCheck.setFlags(item.flags() & ~QtCore.Qt.ItemIsEditable \
                                                   & ~QtCore.Qt.ItemIsDropEnabled\
                                                   & ~QtCore.Qt.ItemIsDragEnabled)
                        lastItem.setFlags(item.flags() & ~QtCore.Qt.ItemIsEditable \
                                                   & ~QtCore.Qt.ItemIsDropEnabled\
                                                   & ~QtCore.Qt.ItemIsDragEnabled)
                        durationItem.setFlags(item.flags() & QtCore.Qt.ItemIsEditable
                                                   & ~QtCore.Qt.ItemIsDropEnabled\
                                                   & ~QtCore.Qt.ItemIsDragEnabled)
                    parentItem.appendRow([item, item2, itemCheck, lastItem, durationItem])

                    for it in p.get(next(iter(p))):
                        par = QtGui.QStandardItem(it.get('label'))
                        par.setFlags(item.flags() & QtCore.Qt.ItemIsEditable & ~QtCore.Qt.ItemIsDragEnabled)
                        if isinstance(it.get('value'), list):
                            print(it.get('value'))
                            li = [float(i) if '.' in i else i for i in it.get('value')]
                            values = str(li).strip('[]')
                        else:
                            values = it.get('value')
                        val = QtGui.QStandardItem(values)
                        val.setFlags(item.flags() & ~QtCore.Qt.ItemIsDragEnabled)
                        item.appendRow([par, val])


class ProxyModel(QSortFilterProxyModel):
    def __init__(self, parent=None):
        super(ProxyModel, self).__init__(parent)


class ComboBoxDelegate(QtWidgets.QStyledItemDelegate):
    instances = []
    def __init__(self, owner, choices):
        super().__init__(owner)
        self.pModel = owner.tableModel
        self.treeView = owner
        list = []
        self.editor = dict()
        for item in choices:
            if item[0] !="":
                list.append(item[0])
        self.items = list

    def commit_editor(self):
        editor = self.sender()
        self.commitData.emit(editor)

    def createEditor(self, parent, option, index):
        # self.updateCombo(index)
        row = index.row()
        hasSample = (self.treeView.model.item(index.parent().row(), 0).child(0, 0).text() == "Sample")
        if row == 0 and index.parent().row() >= 0 and hasSample:# not the very first row
            editor = QtWidgets.QComboBox(parent)
            self.editor[index] = editor
            ComboBoxDelegate.instances.append(editor)
            editor.addItems(self.items)
            editor.currentTextChanged.connect(self.commit_editor)
            editor.highlighted.connect(self.commit_editor)
            return editor
        else:
            editor = QtWidgets.QStyledItemDelegate.createEditor(self, parent, option, index)
            # editor.model().select()
            return editor

    def paint(self, painter, option, index):
        # self.updateCombo(index)
        try:
            hasSample = (self.treeView.model.item(index.row(), 0).child(0, 0).text() == "Sample")
        except AttributeError:
            pass
        if index.column() == 1 and index.parent().row() >= 0 and index.row() == 0 and hasSample:
            value = index.data(QtCore.Qt.DisplayRole)
            style = QtWidgets.QApplication.style()
            opt = QtWidgets.QStyleOptionComboBox()
            opt.text = str(value)
            opt.rect = option.rect
            style.drawComplexControl(QtWidgets.QStyle.CC_ComboBox, opt, painter)
            QtWidgets.QStyledItemDelegate.paint(self, painter, option, index)
            self.treeView.openPersistentEditor(index)
        else:
            QtWidgets.QStyledItemDelegate.paint(self, painter, option, index)

    def setModelData(self, editor, model, index):
        self.updateCombo(index)
        row = index.row()
        hasSample = (model.item(index.parent().row(), 0).child(0, 0).text() == "Sample")
        if row == 0 and index.parent().row() >= 0 and hasSample:
            if model.item(index.parent().row(), 0).child(0, 0).text() == "Sample":
                model.setData(index, editor.currentText(), QtCore.Qt.DisplayRole)
                editor.setCurrentIndex(editor.findText(editor.currentText()))
        else:
            QtWidgets.QStyledItemDelegate.setModelData(self, editor, model, index)

    def updateEditorGeometry(self, editor, option, index):
        editor.setGeometry(option.rect)

    def updateCombo(self, index):
        # print(self.parent().itemDelegateForColumn(1))
        if isinstance(self.sender(), TableModel):
            indToDelete = []
            for ind in self.editor:
                try:
                    currentIndex = self.editor[ind].findText(self.editor[ind].currentText())
                except RuntimeError:
                    #print("oops")
                    indToDelete.append(ind)
                else:
                    pass
                    # print(currentIndex)
            # print(indToDelete)
            for ind in indToDelete:
                del self.editor[ind]
            for ind in self.editor:
                currentIndex = self.editor[ind].findText(self.editor[ind].currentText())
                self.editor[ind].clear()
                list = []
                for item in self.pModel._data:
                    if item[0] != "":
                        list.append(item[0])
                self.items = list
                self.editor[ind].addItems(self.items)
                self.editor[ind].setCurrentIndex(currentIndex)


class ProgressDelegate(QtWidgets.QStyledItemDelegate):
    def paint(self, painter, option, index):
        if index.parent().isValid():
            view = option.widget
            if isinstance(view, QtWidgets.QTreeView) and index.model() is view.model():
                view.openPersistentEditor(index)
            return
        super(ProgressDelegate, self).paint(painter, option, index)

    def createEditor(self, parent, option, index):
        if index.parent().isValid():
            editor = QtWidgets.QProgressBar(parent)
            editor.setFixedWidth(100)
            editor.setContentsMargins(0, 0, 0, 0)
            editor.setValue(45)#index.data(Qt.UserRole))
            return editor
        super(ProgressDelegate, self).createEditor(parent, option, index)


class SpinBoxDelegate(QtWidgets.QStyledItemDelegate):
    def __init__(self, min, max, intOnly=False, parent=None):
        super(SpinBoxDelegate, self).__init__(parent)
        self.min = min
        self.max = max
        self.int = intOnly

    def createEditor(self, parent, option, index):
        if self.int:
            spinBox = QSpinBox(parent)
        else:
            spinBox = QDoubleSpinBox(parent)
            spinBox.setDecimals(3)
        spinBox.setMinimum(self.min)
        spinBox.setMaximum(self.max)
        return spinBox

    def setEditorData(self, spinBox, index):
        # print(index.model().data(index, QtCore.Qt.EditRole))
        value = index.model().data(index, QtCore.Qt.EditRole)
        try:
            spinBox.setValue(float(value))
        except ValueError:
            spinBox.setValue(0.0)

    def setModelData(self, spinBox, model, index):
        spinBox.interpretText()
        value = spinBox.value()
        model.setData(index, value, QtCore.Qt.EditRole)


class TableModel(QtCore.QAbstractTableModel):

    def __init__(self, view, data, headerData):
        super(TableModel, self).__init__()
        self.view = view
        self._data = data
        self.header = headerData
        self.dataChanged.connect(self.onSampleChange)

    def getData(self):
        return self._data

    def data(self, index, role):
        if role == Qt.DisplayRole or role == Qt.EditRole:
            value = self._data[index.row()][index.column()]
            return str(value)

    def setData(self, index, value, role=QtCore.Qt.EditRole):
        if not index.isValid():
            return False
        if role == QtCore.Qt.EditRole:
            self._data[index.row()][index.column()] = str(value)
            self.dataChanged.emit(index, index)#, [QtCore.Qt.DisplayRole])
            return True
        else:
            return False
        return QtCore.QAbstractTableModel.setData(self, index, value, role)# True

    def rowCount(self, index):
        return len(self._data)
        #return self._data.shape[0]

    def columnCount(self, index):
        return len(self._data[0])
        # return self._data.shape[1]

    def flags(self, index):
        return QtCore.Qt.ItemIsEditable | QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable

    def onSampleChange(self):
        for row in range(self.view.model.rowCount()):
            if self.view.model.item(row, 0).child(0, 0).text() == "Sample":
                # print(self.view.model.item(row, 0).child(0, 1).updateCombo())
                self.view.update_summary()

    def update_model(self, newdata):
        self.beginResetModel()
        self._data = newdata
        self.view.update_summary()
        self.endResetModel()


    def headerData(self, section, orientation, role):
        if orientation == Qt.Horizontal and role == Qt.DisplayRole:
            return self.header[section]
        if orientation == Qt.Vertical and role == Qt.DisplayRole:
            return str(section+1)

class Tree(QtWidgets.QTreeView):
    def __init__(self, parent=None):
        super(Tree, self).__init__(parent)
        
        self.model = myStandardItemModel()
        self.setModel(self.model)

        # DATA
        # sampleTable = np.array([["S12345", "11", "12", "13", "14"],
        #                         ["S2", "21", "22", "23", "24"],
        #                         ["S3", "31", "32", "33", "34"], ])

        self.sampleTable = [["S1", "100", "1", "1", "1", "1", "1", "1", "1"],
                       ["S2", "200", "2", "2", "2", "2", "2", "2", "2"],
                       ["S3", "300", "3", "3", "3", "3", "3", "3", "3"],
                       ["S4", "400", "4", "4", "4", "4", "4", "4", "4"],
                       ["", "", " ", " ", " ", " ", " ", " ", " "],
                       ["", "", " ", " ", " ", " ", " ", " ", " "],
                       ["", "", " ", " ", " ", " ", " ", " ", " "]]

        self.headerLabels_sampTable = ["Sample/Title", "Trans", "Height", "Phi Offset",\
                                       "Psi", "Footprint", "Resolution", "Coarse_noMirror", "Switch"]
        self.tableModel = TableModel(self, self.sampleTable, self.headerLabels_sampTable)

        self.delegate = ComboBoxDelegate(self, self.sampleTable)
        self.setItemDelegateForColumn(1, self.delegate)
        self.tableModel.dataChanged.connect(self.delegate.updateCombo)
        self.tableModel.layoutChanged.connect(self.delegate.updateCombo)
        self.tableModel.dataChanged.connect(self.update_summary)
        self.model.dataChanged.connect(self.update_runtime)

        self.dur_delegate = ProgressDelegate(self)
        self.setItemDelegateForColumn(4, self.dur_delegate)

        #self.setEditTriggers(QtWidgets.QAbstractItemView.AllEditTriggers)#CurrentChanged)
        self.setDragDropMode(QtWidgets.QAbstractItemView.InternalMove | QtWidgets.QAbstractItemView.DragDrop)
        self.setDragEnabled(True)
        self.setAcceptDrops(True)

        self.setSelectionMode(QtWidgets.QAbstractItemView.ExtendedSelection)
        self.resizeColumnToContents(0)        
        self.setAlternatingRowColors(True)
        font = QtGui.QFont("Verdana", 10)
        font.setWeight(QtGui.QFont.Bold)
        self.header().setFont(font)
        self.resize(self.sizeHint().height(), self.minimumHeight())
        
        self.setContextMenuPolicy(Qt.CustomContextMenu)
        self.customContextMenuRequested.connect(self.open_menu)
        
        self.collapsed.connect(self.show_summary)
        self.expanded.connect(self.remove_summary)

        self.shortcut = QShortcut(QKeySequence("Del"), self)
        self.shortcut.activated.connect(self.del_action)

        self.update_summary()

        # define shortcuts
        self.menu = QtWidgets.QMenu()
        self.sub_menu = QtWidgets.QMenu("Insert Action")
        self.menu.addMenu(self.sub_menu)
        # actions = SANSActions.actions
        actions = [cls.__name__ for cls in SANSActions.ScriptActionClass.ActionClass.__subclasses__()]
        for name in actions:
            shortcut = "Ctrl+" + name[0].lower()
            action = self.sub_menu.addAction(name)
            #action.setShortcut(QtGui.QKeySequence(shortcut))
            short = QtWidgets.QShortcut(QtGui.QKeySequence(shortcut), self)
            short.activated.connect(partial(self.menu_action, action))
            action.triggered.connect(partial(self.menu_action, action))

    def update_sample_table(self):
        self.tableModel.update_model(self.sampleTable)


    def update_runtime(self):
        self.totalTime = 0.0
        global myActions
        for row in range(self.model.rowCount()):
            try:
                it = self.model.getRootItem(row, 0)
                clName = it.text()
                MyClass = getattr(importlib.import_module(myActions), clName)
                # Instantiate the class (pass arguments to the constructor, if needed)
                tempAction = MyClass()
                tempAction.makeAction(it)
                it.setToolTip(tempAction.toolTip())#
                self.model.item(row, 3).setText(str(row + 1))

            except AttributeError as err:
                # logging.warning("makeAction method undefined in "+tempAction.__repr__)
                print("makeAction method undefined in ", tempAction, err)
            try:
                self.rtime = tempAction.calcTime(self.parent().parent().instrumentSelector.currentText())
                self.model.item(row, 4).setText(str(self.rtime))
                self.totalTime += self.rtime
            except AttributeError:
                pass
        self.parent().parent().runTime.setText("{:02d}h {:02d}min".format(int(divmod(self.totalTime, 60)[0]), int(divmod(self.totalTime, 60)[1])))
        # update window title to indicate unsaved changes:
        self.parent().parent().parent().setWindowModified(True)

    def update_summary(self):
        print(self.model.rowCount())
        for row in range(self.model.rowCount()):
            self.show_summary(self.model.index(row, 0))


    def show_summary(self, index):
        it = self.model.getRootItem(index.row(), index.column())
        global SANSActions
        global myActions
        importlib.reload(SANSActions)
        print("Here:", SANSActions, "MyActions: ", myActions)
        colors = ["black", "red", "blue", "green", "orange", "darkviolet", "salmon", "lavender"]

        clName = it.text()
        MyClass = getattr(importlib.import_module(myActions), clName)
        # Instantiate the class (pass arguments to the constructor, if needed)
        tempAction = MyClass()

        try:
            tempAction.makeAction(it)
        except AttributeError as err:
            print("ShowSummary: makeAction method undefined in ", tempAction)

        if len(self.delegate.items) and it.child(0, 1).text() in self.delegate.items:
            colorIndex = self.delegate.items.index(it.child(0, 1).text())
            self.model.invisibleRootItem().child(index.row(), 1).\
                        setForeground(QtGui.QColor(QtGui.QColor(colors[colorIndex])))
        
        self.model.invisibleRootItem().child(index.row(), 1).setText(tempAction.summary())
        self.resizeColumnToContents(1)


        try:
            if tempAction.isValid()[0]:
                self.model.invisibleRootItem().child(index.row(), 2).\
                            setBackground(QtGui.QColor(QtGui.QColor("green")))
                item = self.model.getRootItem(index.row(), 2)
                item.setToolTip("")
                # self.update_runtime()
            else:
                self.model.invisibleRootItem().child(index.row(),2).\
                            setBackground(QtGui.QColor(QtGui.QColor("red")))
                item = self.model.getRootItem(index.row(), 2)
                item.setToolTip(tempAction.isValid()[1])

        except AttributeError as err:
            print("isValid method undefined in ", tempAction)

        #return self.rtime, tempAction.summary()  # shortText
        # self.update_runtime()
        del tempAction
        
    def remove_summary(self, index):
        self.model.invisibleRootItem().child(index.row(), 1).setText("")
 
        
    def open_menu(self, position):
        # self.menu = QtWidgets.QMenu()
        index = self.indexAt(position)
         
        # self.sub_menu = QtWidgets.QMenu("Insert Action")
        # self.menu.addMenu(self.sub_menu)

        deleteAction = QtWidgets.QAction(QtGui.QIcon('exit.png'), '&Delete row(s)', )
        #deleteAction = self.menu.addAction("Delete Row(s)", self.del_action)
        deleteAction.setShortcut(QKeySequence("Del"))
        deleteAction.triggered.connect(self.del_action)
        self.menu.addAction(deleteAction)
        action = self.menu.exec_(self.viewport().mapToGlobal(position))
        
        if action == deleteAction:
            pass
            #for items in self.selectionModel().selectedRows():
             #   self.model.takeRow(items.row())


    def del_action(self):
        # for items in self.selectionModel().selectedRows():
        #     print(self.model.item(items.row(), 0).child(0,1))
        #     self.model.removeRow(items.row())

        index_list = []
        for model_index in self.selectionModel().selectedRows():
            index = QtCore.QPersistentModelIndex(model_index)
            index_list.append(index)

        for index in index_list:
            self.model.removeRow(index.row())


    def menu_action(self, action):
        global myActions
        item2 = QtGui.QStandardItem("")
        item = QtGui.QStandardItem(action.text())
        itemCheck = QtGui.QStandardItem("")
        rowNumberItem = QtGui.QStandardItem("")
        durationItem = QtGui.QStandardItem("")
        # override index temporarily test
        selection = self.selectedIndexes()
        if selection:
            index = self.selectedIndexes()[0]
            row = index.row()
        else:
            row = -1

        clName = item.text()
        MyClass = getattr(importlib.import_module(myActions), clName)
        # Instantiate the class (pass arguments to the constructor, if needed)
        actionItem = MyClass()

        item.setFlags(item.flags() & ~QtCore.Qt.ItemIsEditable \
                                  & ~QtCore.Qt.ItemIsDropEnabled\
                                  & ~QtCore.Qt.ItemIsDragEnabled)
        root = self.model.invisibleRootItem()
        root.insertRow(row+1, [item, item2, itemCheck, rowNumberItem, durationItem])
        newRow = self.model.item(row+1)
        for key in actionItem.__dict__:
            par = QtGui.QStandardItem(key)
            par.setFlags(newRow.flags() & QtCore.Qt.ItemIsEditable & ~QtCore.Qt.ItemIsDragEnabled)
            if isinstance(actionItem.__dict__[key], list):
                values = str(actionItem.__dict__[key]).strip('[]')
            else:
                values =str(actionItem.__dict__[key])

            val = QtGui.QStandardItem(values)
            newRow.appendRow([par, val])
            self.resizeColumnToContents(0)
        del actionItem
        
    def on_context_menu_factions(self, pos):
        self.menu.exec_( QtGui.QCursor.pos() )

       

            
##### Main App ###########
                    
class App(QtWidgets.QWidget):
    def __init__(self, parent=None):
        super(App, self).__init__(parent)

        global SANSActions
        global myActions
        self.view = Tree(self)
        
        rows = self.view.model.rowCount()
        for i in range(rows):
            self.view.show_summary(self.view.model.index(i, 0))
            
        # set the font
        myFont = "Verdana" # Consolas ok
        font = QtGui.QFont(myFont, 10)
        self.view.setFont(font)

        self.dataGroupBox = QtWidgets.QGroupBox("NR Script")
        dataLayout = QtWidgets.QHBoxLayout()
               

        #dataLayout.addWidget(actionList)
        dataLayout.addWidget(self.view)
        
        self.dataGroupBox.setLayout(dataLayout)
        
        mainLayout = QtWidgets.QVBoxLayout()

        self.instrumentSelector = QtWidgets.QComboBox()
        self.instrumentSelector.addItems(SANSActions.instruments)
        self.instrumentSelector.setFont(QtGui.QFont(myFont, 12, QtGui.QFont.Black))

        buttonLayout = QtWidgets.QHBoxLayout()
        buttonLayout.addWidget(self.instrumentSelector)
        self.printTreeButton = QtWidgets.QPushButton("Output script (Ctrl+p)", self)
        self.printTreeButton.setFont(QtGui.QFont(myFont, 12, QtGui.QFont.Black))
        self.printTreeButton.setShortcut('Ctrl+p')
        self.printTreeButton.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self.printTreeButton.clicked.connect(self.on_print_tree)
        buttonLayout.addWidget(self.printTreeButton)

        self.playButton = QtWidgets.QPushButton("", self)
        self.playButton.setFont(QtGui.QFont(myFont, 12, QtGui.QFont.Black))
        self.playButton.setShortcut('Ctrl+r')
        self.playButton.setIcon(QtGui.QIcon('Icons/play_icon.png'))
        self.playButton.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self.playButton.clicked.connect(self.play_script)
        buttonLayout.addWidget(self.playButton)

        self.pauseButton = QtWidgets.QPushButton("", self)
        self.pauseButton.setFont(QtGui.QFont(myFont, 12, QtGui.QFont.Black))
        self.pauseButton.setShortcut('Ctrl+/')
        self.pauseButton.setIcon(QtGui.QIcon('Icons/pause_icon.png'))
        self.pauseButton.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self.pauseButton.clicked.connect(self.pause_script)
        buttonLayout.addWidget(self.pauseButton)

        self.stopButton = QtWidgets.QPushButton("", self)
        self.stopButton.setFont(QtGui.QFont(myFont, 12, QtGui.QFont.Black))
        self.stopButton.setShortcut('Ctrl+.')
        self.stopButton.setIcon(QtGui.QIcon('Icons/stop_icon.png'))
        self.stopButton.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self.stopButton.clicked.connect(self.stop_script)
        buttonLayout.addWidget(self.stopButton)

        self.printSamplesButton = QtWidgets.QPushButton("Print samples")
        self.printSamplesButton.clicked.connect(self.on_print_samples)
        # buttonLayout.addWidget(self.printSamplesButton)
        self.timeLabel = QLabel("Duration:~ ")
        self.timeLabel.setAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)
        self.timeLabel.setFont(QtGui.QFont(myFont, 12, QtGui.QFont.Black))
        self.runTime = QLabel("00:00")
        self.runTime.setFont(QtGui.QFont(myFont, 12, QtGui.QFont.Black))
        buttonLayout.addWidget(self.timeLabel)
        buttonLayout.addWidget(self.runTime)

        self.actionsLabel = QtWidgets.QLabel("Actions file: ")
        self.actionsEdit = QtWidgets.QLineEdit()
        self.actionsOpenButton = QtWidgets.QPushButton("...")
        self.actionsOpenButton.clicked.connect(self.on_refresh_actions)

        self.fileLabel = QtWidgets.QLabel("Script file: ")
        self.fileEdit = QtWidgets.QLineEdit()
        self.fileOpenButton = QtWidgets.QPushButton("...")
        self.fileOpenButton.clicked.connect(self.on_open_file)

        fileLayout = QtWidgets.QHBoxLayout()
        fileLayout.addWidget(self.actionsLabel)
        fileLayout.addWidget(self.actionsEdit)
        fileLayout.addWidget(self.actionsOpenButton)
        fileLayout.addWidget(self.fileLabel)
        fileLayout.addWidget(self.fileEdit)
        fileLayout.addWidget(self.fileOpenButton)
        
        mainLayout.addLayout(fileLayout)
        mainLayout.addLayout(buttonLayout)

        # Sample Table - this is NR specific and should be moved to SANSActions.py!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
        self.sampleTableGroupBox = QtWidgets.QGroupBox("Sample table")
        sampleLayout = QtWidgets.QHBoxLayout()

        self.sampleTable = QtWidgets.QTableView()
        # Delegates for entering correct numbers --- THIS IS NR SPECIFIC ##################################
        self.transDelegate = SpinBoxDelegate(-600.0, 600.0)
        self.sampleTable.setItemDelegateForColumn(1, self.transDelegate)

        self.heightDelegate = SpinBoxDelegate(-30.0, 30.0)
        self.sampleTable.setItemDelegateForColumn(2, self.heightDelegate)

        self.phiDelegate = SpinBoxDelegate(-5.0, 5.0)
        self.sampleTable.setItemDelegateForColumn(3, self.phiDelegate)

        self.psiDelegate = SpinBoxDelegate(-5.0, 5.0)
        self.sampleTable.setItemDelegateForColumn(4, self.psiDelegate)

        self.fpDelegate = SpinBoxDelegate(1.0, 300.0)
        self.sampleTable.setItemDelegateForColumn(5, self.fpDelegate)

        self.resDelegate = SpinBoxDelegate(0.005, 0.15)
        self.sampleTable.setItemDelegateForColumn(6, self.resDelegate)

        self.coarseNoMDelegate = SpinBoxDelegate(-60, 100.0)
        self.sampleTable.setItemDelegateForColumn(7, self.coarseNoMDelegate)

        self.switchDelegate = SpinBoxDelegate(1.0, 6.0, True)
        self.sampleTable.setItemDelegateForColumn(8, self.switchDelegate)

        ##################################################################################################
        # tableModel = TableModel(sampleTable)
        self.sampleTable.setAlternatingRowColors(True)
        self.sampleTable.setModel(self.view.tableModel)
        self.sampleTable.resizeColumnsToContents()

        # Create buttons:
        self.addLayerButton = QtWidgets.QPushButton("Add sample")
        self.addLayerButton.clicked.connect(self.on_addLayer)
        self.removeLayerButton = QtWidgets.QPushButton("Remove Sample(s)")
        self.removeLayerButton.clicked.connect(self.on_removeLayer)

        self.hbox_layers = QtWidgets.QHBoxLayout()

        self.hbox_layers.addWidget(self.addLayerButton)
        self.hbox_layers.addWidget(self.removeLayerButton)
        self.hbox_layers.addStretch(1)
        self.vbox_samples = QtWidgets.QVBoxLayout()

        # checkbox to hide table
        self.b = QtWidgets.QCheckBox("Hide/Show Sample Table", self)
        self.b.stateChanged.connect(self.click_box)
        mainLayout.addWidget(self.dataGroupBox)
        mainLayout.addWidget(self.sampleTableGroupBox)

        self.vbox_samples.addWidget(self.b)
        self.vbox_samples.addLayout(self.hbox_layers)
        self.vbox_samples.addWidget(self.sampleTable)
        sampleLayout.addLayout(self.vbox_samples)
        self.sampleTableGroupBox.setLayout(sampleLayout)
        self.setLayout(mainLayout)
        
        self.view.resizeColumnToContents(0)
        self.view.resizeColumnToContents(2)
        self.view.header().resizeSection(2, 5)
        
        self.show()
        
###########
    def play_script(self):
        self.view.model.getRootItem(0, 0).setBackground(QtGui.QBrush(Qt.BDiagPattern))
        self.view.model.getRootItem(0, 1).setBackground(QtGui.QBrush(Qt.BDiagPattern))
        self.playButton.setDisabled(1)
        self.pauseButton.setEnabled(1)

    def pause_script(self):
        self.view.model.getRootItem(0, 0).setBackground(QtGui.QBrush(Qt.BDiagPattern))
        self.view.model.getRootItem(0, 1).setBackground(QtGui.QBrush(Qt.BDiagPattern))
        if not self.playButton.isEnabled() and self.pauseButton.isEnabled():
            self.pauseButton.setDisabled(1)
            self.playButton.setEnabled(1)



    def stop_script(self):
        self.view.model.getRootItem(0, 0).setBackground(QtGui.QBrush(Qt.NoBrush))
        self.view.model.getRootItem(0, 1).setBackground(QtGui.QBrush(Qt.NoBrush))
        self.playButton.setDisabled(0)
        self.pauseButton.setDisabled(0)

    def on_refresh_actions(self):
        print(self.actionsEdit.text())
        global myActions
        global SANSActions
        # del myActions
        SANSActions = importlib.import_module(self.actionsEdit.text())
        importlib.reload(SANSActions)
        myActions = self.actionsEdit.text()
        print(SANSActions)
        print(myActions)
        
    def on_open_file(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        fileName, _ = QFileDialog.getSaveFileName(self,"Open Script File...", "",\
                                                  "Python Files (*.py);;All Files (*);;OpenGenie Files (*.gcl)", options=options)

        try:
            with open(fileName) as f:
                self.fileEdit.setText(fileName)
                return fileName
        except IOError:
            print("Created new file " + fileName)
            self.fileEdit.setText(fileName)
            f = open(fileName, "w")
            f.close()

        # if fileName:
        #     self.fileEdit.setText(fileName)
        #     return fileName
        
    def on_print_tree(self):
        self.view.update_summary()
        if os.path.isfile(self.fileEdit.text()):
            f = open(self.fileEdit.text(),"w+")
        else:
            f = self.on_open_file()
            if os.path.isfile(self.fileEdit.text()):
                f = open(self.fileEdit.text(), "w+")
            else:
                return

        # NEED TO REVISIT THIS
        args = []
        for col in range(self.view.tableModel.columnCount(QModelIndex)):
            args.append(0)
            # if self.sampleTable.item(0,col).checkState()  == Qt.Checked:
            #     args.append(1)
            # else:
            #     args.append(0)
        f.write(SANSActions.writeHeader(self.on_print_samples(), args))


        for row in range(self.view.model.rowCount()):
            it = self.view.model.item(row)
            MyClass = getattr(importlib.import_module(myActions), it.text())
           # Instantiate the class (pass arguments to the constructor, if needed)
            tempAction = MyClass()
            tempAction.makeAction(it)
            # Check if sample entry exists for Action:
            if it.child(0, 0).text() == "Sample":
                sampleNumber = self.view.delegate.items.index(it.child(0, 1).text()) + 1
            else:
                sampleNumber = -1
            f.write(tempAction.stringLine(sampleNumber)+"\n")

        f.write(SANSActions.writeFooter(self.on_print_samples()))
        f.close()

    def on_print_samples(self):
        samples = []
        #print(self.view.tableModel.getData())
        for row in range(self.view.tableModel.rowCount(QModelIndex)):
            # print(self.view.tableModel.getData()[row][0])
            #print(self.view.tableModel.headerData(0, Qt.Horizontal, Qt.DisplayRole))

            if self.view.tableModel.getData()[row][0] != "":
                sampleDict = {}
                for col in range(self.view.tableModel.columnCount(QModelIndex)):
                    sampleDict[self.view.tableModel.headerData(col, Qt.Horizontal, Qt.DisplayRole)] = self.view.tableModel.getData()[row][col]
                    # print(sampleDict)
                samples.append(sampleDict)
        return samples

   
    def click_box(self, state):
        if state == QtCore.Qt.Checked:
            self.addLayerButton.setHidden(True)
            self.removeLayerButton.setHidden(True)
            self.sampleTable.setHidden(True)
        else:
            self.addLayerButton.setHidden(False)
            self.removeLayerButton.setHidden(False)
            self.sampleTable.setHidden(False)
    
    def on_addLayer(self):
        pass
        # data = self.view.tableModel.getData()
        # rowNo = len(data)
        # data.append(["S"+str(rowNo+1), "", "", "", "", "", "", "", ""])
        # self.view.tableModel = TableModel(data, self.view.headerLabels_sampTable)
        # self.sampleTable.setModel(self.view.tableModel)
        #
        # self.view.tableModel.rowsInserted.connect(self.view.delegate.updateCombo)


    def on_removeLayer(self):
        pass
        

class MyMainWindow(QtWidgets.QMainWindow):

    def __init__(self, parent=None):
        
        super(MyMainWindow, self).__init__(parent)
        self.form_widget = App(self) 
        self.setCentralWidget(self.form_widget)    
        
        openAct = QtWidgets.QAction(QtGui.QIcon('exit.png'), '&Open...', self)        
        openAct.setShortcut('Ctrl+O')
        openAct.setStatusTip('Open script')
        openAct.triggered.connect(self.openScript)

        saveAct = QtWidgets.QAction(QtGui.QIcon('exit.png'), '&Save as...', self)
        saveAct.setShortcut('Ctrl+S')
        saveAct.setStatusTip('Save script')
        saveAct.triggered.connect(self.saveScript)
        
        exitAct = QtWidgets.QAction(QtGui.QIcon('exit.png'), '&Exit', self)        
        exitAct.setShortcut('Ctrl+Q')
        exitAct.setStatusTip('Exit application')
        exitAct.triggered.connect(qApp.quit)

        change_subtitle_Act = QtWidgets.QAction(QtGui.QIcon('exit.png'), 'E&dit subtitles', self)
        change_subtitle_Act.setShortcut('Ctrl+E')
        change_subtitle_Act.setStatusTip('Edit subtitles')
        change_subtitle_Act.triggered.connect(self.edit_subtitles)

        self.statusBar()

        menubar = self.menuBar()
        fileMenu = menubar.addMenu('&File')
        fileMenu.addAction(openAct)
        fileMenu.addAction(saveAct)
        fileMenu.addAction(exitAct)

        editMenu = menubar.addMenu('&Edit')
        editMenu.addAction(change_subtitle_Act)

        QtWidgets.qApp.installEventFilter(self)

        # Open default file
        fileName = "MARI_test.json"  #"INTER_4Samples_2Contrasts_inject.json" #"Muon_test.json"
        self.form_widget.view.model.populate(fileName)
        for col in range(self.form_widget.view.model.columnCount()):
            self.form_widget.view.resizeColumnToContents(col)

        with open(fileName) as json_file:
            data = json.load(json_file)
            """
            for row, samp in enumerate(data['Samples'], start=1):
                if len(samp):
                    for col in range(self.form_widget.sampleTable.columnCount()):
                        self.form_widget.sampleTable.item(row, col).setText(samp[col])
            """
        rows = self.form_widget.view.model.rowCount()
        for i in range(rows):
            if self.form_widget.view.model.item(i, 0).child(0, 0).text() == "Sample":
                self.form_widget.view.openPersistentEditor(self.form_widget.view.model.item(i, 0).child(0, 1).index())
            self.form_widget.view.show_summary(self.form_widget.view.model.index(i, 0))
        self.form_widget.view.resizeColumnToContents(0)
        self.form_widget.fileEdit.setText('PyScript_test1.py')
        ############################
        self.setStyleSheet("""

            QTreeView {
            alternate-background-color: #e8f4fc;
            background: #f6fafb;
            }

        """)
        
    def openScript(self):
        print("Opening...")
        buttonReply = QMessageBox.question(self, 'MaxSkript', "All changes will be lost. Do you want to save?",\
                                           QMessageBox.Yes | QMessageBox.No | QMessageBox.Cancel, QMessageBox.Cancel)
        if buttonReply == QMessageBox.Yes:
            print('Yes clicked.')
            self.saveScript()
        if buttonReply == QMessageBox.No:
            print('No clicked.')
        if buttonReply == QMessageBox.Cancel:
            return
        
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        fileName, _ = QFileDialog.getOpenFileName(self, "Open saved ScriptMaker state", "",\
                                                  "MaxScript Files (*.json);;"
                                                  "SANS Excel table (*.xlsx);;"
                                                 "All Files (*)", options=options)
        if fileName:
            if fileName.split(".")[1].upper() == "JSON":
                # delete all rows
                self.form_widget.view.model.removeRows(0, self.form_widget.view.model.rowCount())

                self.form_widget.view.model.populate(fileName)
                with open(fileName) as json_file:
                    data = json.load(json_file)
                    #print(len(data['Samples']))
                    # tableModel = TableModel(self.form_widget.view, data['Samples'], self.form_widget.view.headerLabels_sampTable)
                    self.form_widget.view.tableModel.layoutAboutToBeChanged.emit()
                    self.form_widget.view.sampleTable = data['Samples']
                    self.form_widget.view.update_sample_table()
                    self.form_widget.view.tableModel.layoutChanged.emit()

                rows = self.form_widget.view.model.rowCount()
                for i in range(rows):
                    self.form_widget.view.show_summary(self.form_widget.view.model.index(i, 0))
                self.form_widget.parent().setWindowTitle("Ma_xSkript - " + fileName + "[*]")

            else: # SANS Excel table file
                # print(fileName.split(".")[1].upper())
                json_file_name = fileName.split(".")[0] + ".json"
                print(json_file_name)
                wb = load_workbook(filename=fileName)
                ws = wb.active
                wb_row = 2
                with open(json_file_name, "w+") as f:
                    outString = "{\"Samples\": ["
                    ## get defined samples from table:
                    data = self.form_widget.sampleTable.model().getData()

                    for row in data:
                        r = str([''.join(x) for x in row])
                        r = str(r).replace("\'", "\"")
                        outString += r + ",\n"
                    outString = outString[:-2] + "],\n"

                    outString += "\n\n\"Action\":[\n"
                    while ws.cell(row=wb_row, column=1).value:
                        do_trans = getattr(importlib.import_module(myActions), 'do_trans')
                        # Instantiate the class (pass arguments to the constructor, if needed)
                        tempAction = do_trans(title=str(ws.cell(row=wb_row, column=7).value),
                                             pos=str(ws.cell(row=wb_row, column=1).value),
                                             uamps=ws.cell(row=wb_row, column=2).value,
                                             thickness=ws.cell(row=wb_row, column=9).value)
                        outString += tempAction.makeJSON() + ","
                        del tempAction

                        do_sans = getattr(importlib.import_module(myActions), 'do_sans')
                        # Instantiate the class (pass arguments to the constructor, if needed)
                        tempAction = do_sans(title=str(ws.cell(row=wb_row, column=7).value) + '_' + \
                                                   str(ws.cell(row=wb_row, column=8).value),
                                             pos=str(ws.cell(row=wb_row, column=1).value),
                                             uamps=ws.cell(row=wb_row, column=4).value,
                                             thickness=ws.cell(row=wb_row, column=9).value)
                        outString += tempAction.makeJSON() + ","
                        # print([ws.cell(row=row, column=col).value for col in range(1, 12)])
                        del tempAction
                        wb_row += 1
                    outString = outString[:-1] + "]\n}"
                    f.write(outString)

                # delete all rows
                self.form_widget.view.model.removeRows(0, self.form_widget.view.model.rowCount())

                self.form_widget.view.model.populate(json_file_name)
                with open(json_file_name) as json_file:
                    data = json.load(json_file)
                    #print(len(data['Samples']))
                    # tableModel = TableModel(self.form_widget.view, data['Samples'], self.form_widget.view.headerLabels_sampTable)
                    self.form_widget.view.tableModel.layoutAboutToBeChanged.emit()
                    self.form_widget.view.sampleTable = data['Samples']
                    self.form_widget.view.update_sample_table()
                    self.form_widget.view.tableModel.layoutChanged.emit()



    def saveScript(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        fileName, _ = QFileDialog.getSaveFileName(self,"Save ScriptMaker state...", "",\
                                                  "Ma_xScript Files (*.json);;All Files (*)", options=options)
        if fileName:
            f = open(fileName, "w+")
        else:
            return

        outString = "{\"Samples\": ["
        ## get defined samples from table:
        data = self.form_widget.sampleTable.model().getData()

        for row in data:
            r = str([''.join(x) for x in row])
            r = str(r).replace("\'", "\"")
            outString += r + ",\n"
        outString = outString[:-2] + "],\n"

        outString += "\n\n\"Action\":[\n"
        for row in range(self.form_widget.view.model.invisibleRootItem().rowCount()):
            it = self.form_widget.view.model.item(row)
            MyClass = getattr(importlib.import_module(myActions), it.text())
           # Instantiate the class (pass arguments to the constructor, if needed)
            tempAction = MyClass()
            tempAction.makeAction(it)
            outString += tempAction.makeJSON()
            if row < self.form_widget.view.model.rowCount() - 1:
                outString += ","
            del tempAction
                
        outString += "]\n}"
        f.write(outString)
        f.close()
        print("Saved.")

    def edit_subtitles(self):
        new_subtitle = self.getText()
        for model_index in self.form_widget.view.selectionModel().selectedRows():
            row = model_index.row()
            it = self.form_widget.view.model.item(row)
            if it.child(1, 0).text() == 'Subtitle':
                it.child(1, 1).setText(new_subtitle)
                self.form_widget.view.show_summary(model_index)


    def getText(self):
        text, okPressed = QtWidgets.QInputDialog.getText(self, "Change subtitles", "New subtitle:", QtWidgets.QLineEdit.Normal, "")
        if okPressed and text != '':
            return text




def main():

    app = QApplication([])
    app.aboutToQuit.connect(app.deleteLater)
    # app.setStyle('Plastique')
    # app.setStyleSheet(qdarkstyle.load_stylesheet(qt_api='pyqt5'))
    foo = MyMainWindow()
    foo.resize(700, 800)
    foo.setWindowTitle("Ma_xSkript [*]")
    # foo.setWindowFilePath("fiel C:/Users/ktd43279/Documents/GitHub/PyScriptMaker/ScriptMaker_v0p71.py")
    foo.setWindowIcon(QtGui.QIcon('Icons/squareCogs.gif'))

    foo.show()
    app.exec_()


if __name__ == '__main__':
    main()
