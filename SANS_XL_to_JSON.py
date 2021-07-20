from openpyxl import load_workbook
import json
from PyQt5.QtWidgets import QFileDialog

wb = load_workbook(filename='Script_Sunday_Evening2.xlsx')
ws = wb.active
row = 2
while ws.cell(row=row, column=1).value:
    print([ws.cell(row=row, column=col).value for col in range(1, 12)])
    row += 1


def makeJSON():
    rdict = {"do_sans": [{ "label": "title", "value": str(self.title) },\
                   { "label": "pos", "value": str(self.pos) },\
                   { "label": "uamps", "value": str(self.uamps)},\
                   {"label": "thickness", "value": str(self.thickness)}]}
    return json.dumps(rdict, indent=4)

def saveScript(self):
    options = QFileDialog.Options()
    options |= QFileDialog.DontUseNativeDialog
    fileName, _ = QFileDialog.getSaveFileName(self, "Save ScriptMaker state...", "", \
                                              "Ma_xScript Files (*.json);;All Files (*)", options=options)
    if fileName:
        f = open(fileName, "w+")
    else:
        return

    outString = "{\"Samples\": ["
    ## get defined samples from table:
    data = self.form_widget.sampleTable.model().getData()

    for row in data:
        r = str([''.join(x) for x in row])
        r = str(r).replace("\'", "\"")
        outString += r + ",\n"
    outString = outString[:-2] + "],\n"

    outString += "\n\n\"Action\":[\n"
    for row in range(self.form_widget.view.model.invisibleRootItem().rowCount()):
        it = self.form_widget.view.model.item(row)
        MyClass = getattr(importlib.import_module(myActions), it.text())
        # Instantiate the class (pass arguments to the constructor, if needed)
        tempAction = MyClass()
        tempAction.makeAction(it)
        outString += tempAction.makeJSON()
        if row < self.form_widget.view.model.rowCount() - 1:
            outString += ","
        del tempAction

    outString += "]\n}"
    f.write(outString)
    f.close()
    print("Saved.")